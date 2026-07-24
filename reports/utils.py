import csv
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count, Avg, F, Q
from decimal import Decimal
from datetime import timedelta

# Models
from medicines.models import Medicine, Category
from inventory.models import Inventory
from purchases.models import Purchase, PurchaseItem
from sales.models import Sale, SaleItem
from suppliers.models import Supplier

# openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ReportLab
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

def get_filtered_data(report_type, params):
    today = timezone.now().date()
    
    if report_type == 'sales':
        queryset = Sale.objects.select_related('cashier').all()
        # Invoice number search
        inv = params.get('invoice_number', '').strip()
        if inv:
            queryset = queryset.filter(invoice_number__icontains=inv)
        # Customer search
        cust = params.get('customer', '').strip()
        if cust:
            queryset = queryset.filter(customer_name__icontains=cust)
        # Cashier filter
        cashier_id = params.get('cashier', '').strip()
        if cashier_id:
            queryset = queryset.filter(cashier_id=cashier_id)
        # Payment Method filter
        pm = params.get('payment_method', '').strip()
        if pm:
            queryset = queryset.filter(payment_method=pm)
        # Date range
        start_date = params.get('start_date', '').strip()
        end_date = params.get('end_date', '').strip()
        if start_date:
            queryset = queryset.filter(sale_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(sale_date__lte=end_date)
        return queryset

    elif report_type == 'purchases':
        queryset = Purchase.objects.select_related('supplier', 'created_by').all()
        # Supplier filter
        sup = params.get('supplier', '').strip()
        if sup:
            queryset = queryset.filter(supplier_id=sup)
        # Status filter
        status = params.get('status', '').strip()
        if status:
            queryset = queryset.filter(status=status)
        # Date range
        start_date = params.get('start_date', '').strip()
        end_date = params.get('end_date', '').strip()
        if start_date:
            queryset = queryset.filter(purchase_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(purchase_date__lte=end_date)
        return queryset

    elif report_type == 'inventory':
        queryset = Inventory.objects.select_related('medicine', 'medicine__category').all()
        # Medicine search
        med = params.get('medicine', '').strip()
        if med:
            queryset = queryset.filter(medicine_id=med)
        # Category filter
        cat = params.get('category', '').strip()
        if cat:
            queryset = queryset.filter(medicine__category_id=cat)
        # Supplier filter (retrieve medicines supplied by vendor)
        sup = params.get('supplier', '').strip()
        if sup:
            queryset = queryset.filter(medicine__purchase_items__purchase__supplier_id=sup).distinct()
        return queryset

    elif report_type == 'medicines':
        queryset = Medicine.objects.select_related('category').filter(is_active=True)
        # Category
        cat = params.get('category', '').strip()
        if cat:
            queryset = queryset.filter(category_id=cat)
        # Supplier
        sup = params.get('supplier', '').strip()
        if sup:
            queryset = queryset.filter(purchase_items__purchase__supplier_id=sup).distinct()

        from django.db.models.functions import Coalesce
        queryset = queryset.annotate(
            total_stock=Coalesce(
                Sum('inventory_batches__quantity', filter=Q(inventory_batches__expiry_date__gte=today)),
                0
            )
        )
        return queryset

    elif report_type == 'suppliers':
        queryset = Supplier.objects.all()
        q_name = params.get('supplier_name', '').strip()
        if q_name:
            queryset = queryset.filter(name__icontains=q_name)
        return queryset

    elif report_type == 'low-stock':
        # Annotate total stock quantity and filter by minimum stock level
        from django.db.models.functions import Coalesce
        queryset = Medicine.objects.select_related('category').annotate(
            total_stock=Coalesce(
                Sum('inventory_batches__quantity', filter=Q(inventory_batches__expiry_date__gte=today)),
                0
            )
        ).filter(total_stock__lt=50, total_stock__gt=0, is_active=True).annotate(
            shortage=50 - F('total_stock')
        )
        return queryset

    elif report_type == 'expiry':
        # Show expired and expiring soon (within 30 days)
        limit_date = today + timedelta(days=30)
        queryset = Inventory.objects.select_related('medicine', 'medicine__category').filter(
            expiry_date__lte=limit_date
        )
        return queryset

    return None

def generate_csv(report_type, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{timezone.now().date()}.csv"'
    writer = csv.writer(response)

    if report_type == 'sales':
        writer.writerow(['Invoice Number', 'Customer Name', 'Cashier', 'Sale Date', 'Subtotal', 'Discount', 'Tax', 'Total Amount', 'Payment Method'])
        for sale in queryset:
            writer.writerow([sale.invoice_number, sale.customer_name or 'Walk-in', sale.cashier.username, sale.sale_date, sale.subtotal, sale.discount, sale.tax, sale.total_amount, sale.payment_method])
            
    elif report_type == 'purchases':
        writer.writerow(['PO ID', 'Supplier', 'Invoice Number', 'Purchase Date', 'Total Amount', 'Status', 'Created By'])
        for p in queryset:
            writer.writerow([f"PO-{p.id}", p.supplier.name, p.invoice_number, p.purchase_date, p.total_amount, p.status, p.created_by.username])
            
    elif report_type == 'inventory':
        writer.writerow(['Medicine', 'Category', 'Batch Number', 'Current Stock', 'Purchase Price', 'Selling Price', 'Expiry Date', 'Status'])
        for i in queryset:
            writer.writerow([i.medicine.name, i.medicine.category.name if i.medicine.category else 'N/A', i.batch_no, i.quantity, i.medicine.purchase_price, i.medicine.selling_price, i.expiry_date, i.status])

    elif report_type == 'medicines':
        writer.writerow(['Medicine Name', 'Brand/Generic Name', 'Category', 'Selling Price', 'Purchase Price', 'Total Stock'])
        for m in queryset:
            stock = m.total_stock
            writer.writerow([m.name, m.brand or 'Generic', m.category.name if m.category else 'N/A', m.selling_price, m.purchase_price, stock])

    elif report_type == 'suppliers':
        writer.writerow(['Supplier Name', 'Contact Person', 'Phone', 'Email', 'Total Purchase Value'])
        for s in queryset:
            total_val = Purchase.objects.filter(supplier=s, status='Received').aggregate(total=Sum('total_amount'))['total'] or 0
            writer.writerow([s.name, s.contact_person or 'N/A', s.phone or 'N/A', s.email or 'N/A', total_val])

    elif report_type == 'low-stock':
        writer.writerow(['Medicine', 'Category', 'Current Stock', 'Minimum Stock Level'])
        for m in queryset:
            stock = m.total_stock or 0
            writer.writerow([m.name, m.category.name if m.category else 'N/A', stock, m.minimum_stock_level])

    elif report_type == 'expiry':
        writer.writerow(['Medicine', 'Batch Number', 'Expiry Date', 'Quantity', 'Status'])
        for i in queryset:
            writer.writerow([i.medicine.name, i.batch_no, i.expiry_date, i.quantity, i.status])

    return response

def generate_excel(report_type, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{timezone.now().date()}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Report Summary"

    # Styling
    font_title = Font(name='Arial', size=14, bold=True, color='0D6EFD')
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    if report_type == 'sales':
        ws.append(["Sales Summary Report"])
        ws.cell(1, 1).font = font_title
        ws.append([])
        
        headers = ['Invoice Number', 'Customer Name', 'Cashier', 'Sale Date', 'Subtotal', 'Discount', 'Tax', 'Total Amount', 'Payment Method']
        ws.append(headers)
        
        for sale in queryset:
            ws.append([sale.invoice_number, sale.customer_name or 'Walk-in', sale.cashier.username, str(sale.sale_date), float(sale.subtotal), float(sale.discount), float(sale.tax), float(sale.total_amount), sale.payment_method])

    elif report_type == 'purchases':
        ws.append(["Purchase Orders Report"])
        ws.cell(1, 1).font = font_title
        ws.append([])
        
        headers = ['PO ID', 'Supplier', 'Invoice Number', 'Purchase Date', 'Total Amount', 'Status', 'Created By']
        ws.append(headers)
        
        for p in queryset:
            ws.append([f"PO-{p.id}", p.supplier.name, p.invoice_number, str(p.purchase_date), float(p.total_amount), p.status, p.created_by.username])

    elif report_type == 'inventory':
        ws.append(["Inventory Levels Report"])
        ws.cell(1, 1).font = font_title
        ws.append([])
        
        headers = ['Medicine', 'Category', 'Batch Number', 'Current Stock', 'Purchase Price', 'Selling Price', 'Expiry Date', 'Status']
        ws.append(headers)
        
        for i in queryset:
            ws.append([i.medicine.name, i.medicine.category.name if i.medicine.category else 'N/A', i.batch_no, i.quantity, float(i.medicine.purchase_price), float(i.medicine.selling_price), str(i.expiry_date), i.status])

    elif report_type == 'medicines':
        ws.append(["Medicines Catalogue Report"])
        ws.cell(1, 1).font = font_title
        ws.append([])
        
        headers = ['Medicine Name', 'Brand/Generic Name', 'Category', 'Selling Price', 'Purchase Price', 'Total Stock']
        ws.append(headers)
        
        for m in queryset:
            stock = m.total_stock
            ws.append([m.name, m.brand or 'Generic', m.category.name if m.category else 'N/A', float(m.selling_price), float(m.purchase_price), stock])

    elif report_type == 'suppliers':
        ws.append(["Suppliers Report"])
        ws.cell(1, 1).font = font_title
        ws.append([])
        
        headers = ['Supplier Name', 'Contact Person', 'Phone', 'Email', 'Total Purchase Value']
        ws.append(headers)
        
        for s in queryset:
            total_val = Purchase.objects.filter(supplier=s, status='Received').aggregate(total=Sum('total_amount'))['total'] or 0
            ws.append([s.name, s.contact_person or 'N/A', s.phone or 'N/A', s.email or 'N/A', float(total_val)])

    elif report_type == 'low-stock':
        ws.append(["Low Stock Warnings Report"])
        ws.cell(1, 1).font = font_title
        ws.append([])
        
        headers = ['Medicine', 'Category', 'Current Stock', 'Minimum Stock Level']
        ws.append(headers)
        
        for m in queryset:
            stock = m.total_stock or 0
            ws.append([m.name, m.category.name if m.category else 'N/A', stock, m.minimum_stock_level])

    elif report_type == 'expiry':
        ws.append(["Expiry Monitoring Report"])
        ws.cell(1, 1).font = font_title
        ws.append([])
        
        headers = ['Medicine', 'Batch Number', 'Expiry Date', 'Quantity', 'Status']
        ws.append(headers)
        
        for i in queryset:
            ws.append([i.medicine.name, i.batch_no, str(i.expiry_date), i.quantity, i.status])

    # Apply headers style
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    # Apply cells formatting
    for row in range(4, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border_thin
            if isinstance(cell.value, float):
                cell.number_format = '"Nrs." #,##0.00'
                cell.alignment = align_right
            elif isinstance(cell.value, int):
                cell.alignment = align_right
            else:
                cell.alignment = align_left

    # Set column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(response)
    return response

def generate_pdf_report(report_type, queryset):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{timezone.now().date()}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        textColor=colors.HexColor('#0d6efd'),
        fontSize=20,
        spaceAfter=8
    )
    meta_style = ParagraphStyle(
        'MetaStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#555555')
    )
    table_cell_style = ParagraphStyle('TableCellStyle', parent=styles['Normal'], fontSize=8, leading=10)
    table_header_style = ParagraphStyle('TableHeaderStyle', parent=styles['Normal'], fontSize=8, leading=10, fontName='Helvetica-Bold', textColor=colors.white)

    story.append(Paragraph(f"{report_type.title()} Report - PharmaCare", title_style))
    story.append(Paragraph(f"Generated at: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
    story.append(Spacer(1, 15))

    if report_type == 'sales':
        data = [[Paragraph("Invoice", table_header_style), Paragraph("Customer", table_header_style), Paragraph("Cashier", table_header_style), Paragraph("Date", table_header_style), Paragraph("Total", table_header_style), Paragraph("Method", table_header_style)]]
        for sale in queryset:
            data.append([
                Paragraph(sale.invoice_number, table_cell_style),
                Paragraph(sale.customer_name or 'Walk-in', table_cell_style),
                Paragraph(sale.cashier.username, table_cell_style),
                Paragraph(str(sale.sale_date), table_cell_style),
                Paragraph(f"Nrs. {sale.total_amount}", table_cell_style),
                Paragraph(sale.payment_method, table_cell_style)
            ])
        col_widths = [100, 120, 80, 80, 80, 80]

    elif report_type == 'purchases':
        data = [[Paragraph("PO ID", table_header_style), Paragraph("Supplier", table_header_style), Paragraph("Invoice", table_header_style), Paragraph("Date", table_header_style), Paragraph("Total", table_header_style), Paragraph("Status", table_header_style)]]
        for p in queryset:
            data.append([
                Paragraph(f"PO-{p.id}", table_cell_style),
                Paragraph(p.supplier.name, table_cell_style),
                Paragraph(p.invoice_number, table_cell_style),
                Paragraph(str(p.purchase_date), table_cell_style),
                Paragraph(f"Nrs. {p.total_amount}", table_cell_style),
                Paragraph(p.status, table_cell_style)
            ])
        col_widths = [80, 140, 100, 80, 70, 70]

    elif report_type == 'inventory':
        data = [[Paragraph("Medicine", table_header_style), Paragraph("Category", table_header_style), Paragraph("Batch", table_header_style), Paragraph("Stock", table_header_style), Paragraph("Price", table_header_style), Paragraph("Expiry", table_header_style)]]
        for i in queryset:
            data.append([
                Paragraph(i.medicine.name, table_cell_style),
                Paragraph(i.medicine.category.name if i.medicine.category else 'N/A', table_cell_style),
                Paragraph(i.batch_no, table_cell_style),
                Paragraph(str(i.quantity), table_cell_style),
                Paragraph(f"Nrs. {i.medicine.selling_price}", table_cell_style),
                Paragraph(str(i.expiry_date), table_cell_style)
            ])
        col_widths = [140, 100, 80, 60, 80, 80]

    elif report_type == 'medicines':
        data = [[Paragraph("Medicine Name", table_header_style), Paragraph("Brand/Generic", table_header_style), Paragraph("Category", table_header_style), Paragraph("Retail Price", table_header_style), Paragraph("Stock", table_header_style)]]
        for m in queryset:
            stock = m.total_stock
            data.append([
                Paragraph(m.name, table_cell_style),
                Paragraph(m.brand or 'Generic', table_cell_style),
                Paragraph(m.category.name if m.category else 'N/A', table_cell_style),
                Paragraph(f"Nrs. {m.selling_price}", table_cell_style),
                Paragraph(str(stock), table_cell_style)
            ])
        col_widths = [180, 120, 100, 80, 60]

    elif report_type == 'suppliers':
        data = [[Paragraph("Supplier Name", table_header_style), Paragraph("Contact Person", table_header_style), Paragraph("Phone", table_header_style), Paragraph("Email", table_header_style)]]
        for s in queryset:
            data.append([
                Paragraph(s.name, table_cell_style),
                Paragraph(s.contact_person or 'N/A', table_cell_style),
                Paragraph(s.phone or 'N/A', table_cell_style),
                Paragraph(s.email or 'N/A', table_cell_style)
            ])
        col_widths = [160, 120, 100, 160]

    elif report_type == 'low-stock':
        data = [[Paragraph("Medicine", table_header_style), Paragraph("Category", table_header_style), Paragraph("Current Stock", table_header_style), Paragraph("Minimum Level", table_header_style)]]
        for m in queryset:
            stock = m.total_stock or 0
            data.append([
                Paragraph(m.name, table_cell_style),
                Paragraph(m.category.name if m.category else 'N/A', table_cell_style),
                Paragraph(str(stock), table_cell_style),
                Paragraph(str(m.minimum_stock_level), table_cell_style)
            ])
        col_widths = [200, 140, 100, 100]

    elif report_type == 'expiry':
        data = [[Paragraph("Medicine", table_header_style), Paragraph("Batch", table_header_style), Paragraph("Expiry Date", table_header_style), Paragraph("Stock Quantity", table_header_style), Paragraph("Status", table_header_style)]]
        for i in queryset:
            data.append([
                Paragraph(i.medicine.name, table_cell_style),
                Paragraph(i.batch_no, table_cell_style),
                Paragraph(str(i.expiry_date), table_cell_style),
                Paragraph(str(i.quantity), table_cell_style),
                Paragraph(i.status, table_cell_style)
            ])
        col_widths = [180, 90, 90, 90, 90]

    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d6efd')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(table)
    
    doc.build(story)
    return response
